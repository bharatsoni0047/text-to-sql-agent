# ingestion/documents.py - turn an uploaded PDF, DOCX, XLSX or TXT file into searchable chunks
import io
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from retrieval import search
import config

# splits long text into overlapping pieces at natural break points
splitter = RecursiveCharacterTextSplitter(chunk_size=config.CHUNK_SIZE,
                                          chunk_overlap=config.CHUNK_OVERLAP)

# what this function does: read every sheet of an excel file as one row of text per line
def read_excel(file_bytes):
  workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
  lines = []
  for sheet in workbook.worksheets:
    lines.append(f"Sheet: {sheet.title}")
    for row in sheet.iter_rows(values_only=True):
      lines.append(" | ".join("" if cell is None else str(cell) for cell in row))
  return "\n".join(lines)

# what this function does: pull the plain text out of a pdf, docx, xlsx or txt file
def extract_text(file_name, file_bytes):
  lowered = file_name.lower()
  if lowered.endswith(".pdf"):
    return "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(file_bytes)).pages)
  if lowered.endswith(".docx"):
    return "\n".join(p.text for p in Document(io.BytesIO(file_bytes)).paragraphs)
  if lowered.endswith((".xlsx", ".xlsm")):
    return read_excel(file_bytes)
  return file_bytes.decode("utf-8", errors="ignore")

# what this function does: split a file into chunks and add them to the search index
def add_document(file_name, file_bytes):
  chunks = splitter.split_text(extract_text(file_name, file_bytes))
  search.add_document_chunks(chunks, file_name)
  return len(chunks)
